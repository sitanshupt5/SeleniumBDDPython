# utilities/utils.py
"""
Utility helpers for reading and writing Excel-based test data.

This module centralizes interactions with the ``testdata/LoginData.xlsx`` file
using both ``pandas`` and ``openpyxl``. It is primarily used to fetch login
credentials and manipulate test data sheets.
"""
import os.path
import openpyxl
import pandas as pd

# Resolve absolute project root and Excel path
PROJECT_ROOT: str = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
EXCEL_FILE_PATH: str = os.path.join(PROJECT_ROOT, "testdata", "LoginData.xlsx")


def get_credentials(user_type: str, sheet: str) -> tuple[str | None, str | None]:
    """
    Retrieve username and password for a given user type from the Excel sheet.

    The Excel file must contain columns: ``usertype``, ``username``, ``password``.

    :param user_type: The logical user type identifier (e.g., ``"admin"``, ``"standard_user"``).
    :param sheet: Worksheet name inside the Excel file.
    :returns: Tuple of ``(username, password)`` if found, otherwise ``(None, None)``.
    """
    df = pd.read_excel(EXCEL_FILE_PATH, sheet_name=sheet)
    try:
        user_data = df[df["usertype"] == user_type]
        username = user_data["username"].values[0]
        password = user_data["password"].values[0]
        return username, password
    except IndexError:
        return None, None


def get_row_count(sheet_name: str) -> int:
    """
    Get the total number of rows in a given Excel sheet.

    :param sheet_name: Worksheet name inside the Excel file.
    :returns: Maximum row index (1-based, includes header row).
    """
    workbook = openpyxl.load_workbook(EXCEL_FILE_PATH)
    sheet = workbook[sheet_name]
    return sheet.max_row


def get_column_count(sheet_name: str) -> int:
    """
    Get the total number of columns in a given Excel sheet.

    :param sheet_name: Worksheet name inside the Excel file.
    :returns: Maximum column index (1-based).
    """
    workbook = openpyxl.load_workbook(EXCEL_FILE_PATH)
    sheet = workbook[sheet_name]
    return sheet.max_column


def read_data(sheet_name: str, row_num: int, column_no: int) -> str | int | float | None:
    """
    Read a single cell value from the Excel sheet.

    :param sheet_name: Worksheet name inside the Excel file.
    :param row_num: Row index (1-based).
    :param column_no: Column index (1-based).
    :returns: Cell value, which may be ``str``, ``int``, ``float``, or ``None`` if empty.
    """
    workbook = openpyxl.load_workbook(EXCEL_FILE_PATH)
    sheet = workbook[sheet_name]
    return sheet.cell(row=row_num, column=column_no).value


def write_data(sheet_name: str, row_num: int, column_no: int, data: object) -> None:
    """
    Write a value into a specific cell of the Excel sheet.

    :param sheet_name: Worksheet name inside the Excel file.
    :param row_num: Row index (1-based).
    :param column_no: Column index (1-based).
    :param data: Value to assign (string, number, or any Excel-compatible type).
    :returns: None
    """
    workbook = openpyxl.load_workbook(EXCEL_FILE_PATH)
    sheet = workbook[sheet_name]
    sheet.cell(row=row_num, column=column_no).value = data
    workbook.save(EXCEL_FILE_PATH)
